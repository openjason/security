'''
-- 导出 sec 的数据库结构
CREATE DATABASE IF NOT EXISTS `sec` /*!40100 DEFAULT CHARACTER SET latin1 */;
USE `sec`;

-- 导出  表 sec.d2407 结构
CREATE TABLE IF NOT EXISTS `d2407` (
  `xdate` datetime DEFAULT NULL,
  `price` float DEFAULT NULL,
  `pdiff` float DEFAULT NULL,
  `volume` float DEFAULT NULL,
  `amount` float DEFAULT NULL,
  `bors` char(8) CHARACTER SET latin1 DEFAULT NULL
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_520_ci;

'''
#Author: JasonChan

"""
功能：将Excel数据导入到MySQL数据库
"""
import openpyxl
import pymysql
# Open the workbook and define the worksheet
book = openpyxl.load_workbook("d:\\temp\\sz2047.xlsx")

sheet = book.get_sheet_by_name('sz002407')

#建立一个MySQL连接
database =pymysql.connect (host="localhost", user = "root", passwd = "", db = "sec")

# 获得游标对象, 用于逐行遍历数据库数据
cursor = database.cursor()

# 创建插入SQL语句
query = """INSERT INTO d2047 (xdate, price, pdiff, volume, amount, bors) VALUES (%s, %s, %s, %s, %s)"""

# 创建一个for循环迭代读取xls文件每行数据的, 从第二行开始是要跳过标题
for r in range(1, sheet.nrows):
    xdate      = sheet.cell(r,).value
    price = sheet.cell(r,1).value
    pdiff          = sheet.cell(r,2).value
    volume     = sheet.cell(r,3).value
    amount       = sheet.cell(r,4).value
    bors = sheet.cell(r,5).value

    values = (xdate, price, pdiff, volume, amount, bors)

    print (query)
    pirnt (values)
      # 执行sql语句
   cursor.execute(query, values)

# 关闭游标
cursor.close()

# 提交
database.commit()

# 关闭数据库连接
database.close()

# 打印结果
print ("")
print ("Done! ")
print ("")
columns = str(sheet.ncols)
rows = str(sheet.nrows)
print ("我刚导入了 " + columns + " 列 and " + rows + " 行数据到MySQL!")